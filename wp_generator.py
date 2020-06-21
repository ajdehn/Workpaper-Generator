import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import sys
import os
import tkinter as tk
from tkinter import filedialog

def main():
	# Read input file (status tracker)
	inputWorkbook = readInputFile()
	
	# Create workpapers and folder structure as defined in the status tracker
	ws = inputWorkbook.active
	for row_cells in ws.iter_rows(min_row=2):
		controlObjective = row_cells[0].value
		controlNumber = row_cells[1].value
		controlNumber = str(round(controlNumber, 1))
		controlDescription = row_cells[2].value
		prepareOutputFile(controlObjective, controlNumber, controlDescription)
	
def readInputFile():
	root = tk.Tk()
	root.withdraw()
	inputFileName = filedialog.askopenfilename(title = "** Select Status Tracker **", initialdir=os.getcwd())
	workbook = load_workbook(filename=inputFileName, data_only=True)
	return workbook

def prepareOutputFile(controlObjective, controlNumber, controlDescription):
	# Create directory for controlObjective if it doesn't already exist
	if not os.path.exists(controlObjective):
		os.makedirs(controlObjective)
	
	# Create Output Workpaper using template
	output = load_workbook(filename='Templates/Workpaper_Template.xlsx')
	sumSheet = output['Leadsheet']
	sumSheet["B4"] = controlNumber
	sumSheet["B5"] = controlDescription
	
	filePath = "%s/%s.xlsx" % (controlObjective, controlNumber)
	
	# Save file to specified directory
	output.save(filename=filePath)
	

	
# Main Function
if __name__ == '__main__':
    main()



